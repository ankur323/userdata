from django.shortcuts import HttpResponse, render
import openpyxl
from models import UploadInfo, UploadSetting
from datetime import datetime
import json
from django.views.decorators.csrf import csrf_exempt
import os
import calendar
from utils import norm_unicode_to_ascii
from upload_util import number_format_types, number_format_parse, date_format_types, date_format_parse


# def index(request):
#     return HttpResponse("<p>Hello Django</p>")

# import this from sentieopackages as in fetch_excel_meta
def default_response():
    return {
        'result': [],
        'counter': 0,
        'response': {
            'status': False,
            'msg': []
        }
    }


@csrf_exempt
def index(request):
    if "GET" == request.method:
        return render(request, 'uploadapp/index.html', {})
    else:
        upload = request.FILES.values()[0]
        data = ''
        for chunk in upload.chunks():
            data += chunk
        # todo handle the cases where multiple files with the same name are uploaded
        # either by the same user or different users
        dirpath = 'C:\\Users\\ankurjaiswal\\Desktop\\UserData\\'
        filename = 'copy' + upload.name
        if os.path.exists(dirpath + filename):
            os.remove(dirpath + filename)
        with open(dirpath + filename, 'wb+') as destination:
            destination.write(data)

        # excel_file = request.FILES["excel_file"]

        wb = openpyxl.load_workbook(dirpath + filename)
        print wb.sheetnames

        all_sheets_data = dict()
        for ws in wb.sheetnames:
            print ws
            # todo make it smart enough to remove columns containing no data
            worksheet = wb[ws]
            current_sheet_data = list()
            print worksheet.max_row, worksheet.max_column
            row_count = worksheet.max_row
            col_count = worksheet.max_column

            for i in range(1, row_count):
                row_data = list()
                for j in range(1, col_count):
                    cell = worksheet.cell(i, j)
                    if cell.value and isinstance(cell.value, basestring):
                        row_data.append(cell.value.decode("utf-8"))
                    else:
                        row_data.append('')

                current_sheet_data.append(row_data)

            # for row in worksheet.iter_rows():
            #     row_data = list()
            #     for cell in row:
            #         # row_data.append(str(cell.value))
            #         if cell.value and isinstance(cell.value, basestring):
            #             row_data.append(cell.value.decode("utf-8"))
            #         else:
            #             row_data.append('')
            #
            #     current_sheet_data.append(row_data)
            all_sheets_data[ws] = current_sheet_data

        ui = UploadInfo()
        ui.filename = str(filename)
        ui.status = 'queued'
        ui.datalist = json.dumps(all_sheets_data)
        ui.uploadtime = calendar.timegm(datetime.now().utctimetuple())
        ui.sheetnames = json.dumps(wb.sheetnames)
        ui.save()

        response = default_response()
        response['result'] = dict()
        response['result']['data'] = ui.datalist
        return HttpResponse(json.dumps(response))

        # return render(request, 'myapp/index.html', {"excel_data": excel_data})


def today_is(request):
    now = datetime.now()
    html = "<html><body>Current date and time: {0}</body></html>".format(now)
    return HttpResponse(html)


@csrf_exempt
def get_upload_history(request):
    # todo make this post
    print request.GET.keys()
    result = list()
    ui_list = UploadInfo.objects.all()
    for ui in ui_list:
        cur_row_dict = {key: ui.__dict__[key] for key in ['id', 'filename', 'status', 'uploadtime']}
        result.append(cur_row_dict)
        print cur_row_dict.keys()
    print result

    response = default_response()
    response['result'] = result
    return HttpResponse(json.dumps(response))


@csrf_exempt
def get_upload_data_for_id(request):
    # todo make this post
    print request.GET.keys()

    result = dict()

    id = request.GET.get('id', None)
    print 'aaa', id
    if id:
        ui = UploadInfo.objects.filter(id=id).first()
        if ui:
            result = {key: ui.__dict__[key] for key in
                      ['id', 'filename', 'status', 'uploadtime', 'sheetnames', 'datalist']}
            result['datalist'] = json.loads(result['datalist'])
            result['sheetnames'] = json.loads(result['sheetnames'])
            # result.append(res)

    response = default_response()
    response['result'] = result
    return HttpResponse(json.dumps(response))


@csrf_exempt
def get_user_upload_settings(request):
    result = list()
    ui_list = UploadSetting.objects.all()
    for ui in ui_list:
        cur_row_dict = {key: ui.__dict__[key] for key in ['id', 'setting_name', 'userid', 'worksheet',
                                                          'delimeter', 'default_ownerid', 'header_row',
                                                          'datastart_row']}
        result.append(cur_row_dict)
    response = default_response()
    response['result'] = result
    return HttpResponse(json.dumps(response))


@csrf_exempt
def get_upload_metadata(request):
    result = dict()
    upload_data_type = ['portfolio', 'watchlist', 'custom ticker', 'time series', 'custom financial data']
    upload_identifier_type = ['isin', 'cusip', 'refinitiv ric', 'sentieo ticker', 'bloomberg ticker',
                              'bloomberg openfigi', 'refinitiv openpermid']

    result['upload_data_type'] = upload_data_type
    result['upload_identifier_type'] = upload_identifier_type
    result['upload_numberformat_type'] = number_format_types
    result['upload_dateformat_type'] = date_format_types
    response = default_response()
    response['result'] = result
    return HttpResponse(json.dumps(response))


@csrf_exempt
def map_data(request):
    # data = request.POST.get('data', None)
    user_settings = request.POST.get('user_settings', None)
    if user_settings:
        user_settings = json.loads(user_settings)
        # print user_settings
        setting_id = user_settings.get('setting_id', None)
        worksheet_name = user_settings['worksheet_name']
        delimiter_string = user_settings['delimiter_string']
        default_owner = user_settings['default_owner']
        header_row = user_settings['header_row']
        datastart_row = user_settings['datastart_row']
        upload_data_type = user_settings['upload_data_type']
        upload_identifier_type = user_settings['upload_identifier_type']
        number_format = user_settings['number_format']
        as_of_date = user_settings['as_of_date']
        date_format = user_settings['date_format']
        is_update = user_settings['is_update']
        is_replace = user_settings['is_replace']
        column_to_identifier_map = user_settings['column_to_identifier_map']
        security_id_map = user_settings['security_id_map']
        search_replace_map = user_settings['search_replace_map']
        print setting_id
        print worksheet_name
        print delimiter_string
        print json.loads(default_owner)
        print header_row
        print datastart_row
        print upload_data_type
        print upload_identifier_type
        print number_format
        print as_of_date
        print date_format
        print is_update
        print is_replace
        print column_to_identifier_map
        print security_id_map
        print search_replace_map



    result = {}
    response = default_response()
    response['result'] = result
    return HttpResponse(json.dumps(response))






